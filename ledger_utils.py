"""
Ledger utility module for reading/writing the Redwed ledger Excel file.
"""
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import re
import time
import json

EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'Redwed ledger.xlsx')

# Smart file-based cache to eliminate loading delays
_cache = {}
_cache_mtime = {}

# Month name normalization map
MONTH_NAMES = {
    1: 'January', 2: 'February', 3: 'March', 4: 'April',
    5: 'May', 6: 'June', 7: 'July', 8: 'August',
    9: 'September', 10: 'October', 11: 'November', 12: 'December'
}

MONTH_NAME_TO_NUM = {}
for num, name in MONTH_NAMES.items():
    MONTH_NAME_TO_NUM[name.lower()] = num
    MONTH_NAME_TO_NUM[name.lower()[:3]] = num
# Handle the typo in the Excel
MONTH_NAME_TO_NUM['feburary'] = 2


def _normalize_month(raw):
    """Normalize month marker text from Excel (e.g. 'J\\nA\\nN\\nU\\nA\\nR\\nY' -> 'january')."""
    if not raw:
        return None
    text = str(raw).replace('\n', '').replace('\r', '').strip().lower()
    return MONTH_NAME_TO_NUM.get(text)


def _parse_date(val):
    """Parse a date value from Excel cell."""
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if val and str(val).strip():
        return str(val).strip()
    return ''


def _parse_and_cache(wb, year_sheet, current_mtime):
    cache_key = f'ledger_{year_sheet}'
    json_cache_path = os.path.join(os.path.dirname(__file__), f'{cache_key}.json')
    ws = wb[year_sheet]
    
    entries = []
    current_month = None
    current_month_num = None
    entry_id = 0
    empty_streak = 0  # Track consecutive empty rows for early exit
    
    # Parse amounts safely
    def safe_float(v):
        if v is None or v == '' or v == ' ':
            return 0.0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0.0
    
    MAX_EMPTY_STREAK = 30
    MAX_SCAN_ROW = 500
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=10, max_row=MAX_SCAN_ROW, max_col=9, values_only=False), start=10):
        row_data = [cell.value for cell in row]
        
        month_val = _normalize_month(row_data[0])
        if month_val:
            current_month = MONTH_NAMES[month_val]
            current_month_num = month_val
            empty_streak = 0
        
        name = row_data[2]
        revenue = row_data[4]
        expenditures = row_data[5]
        investment = row_data[6]
        salary = row_data[7]
        
        has_data = name or any(
            v is not None and v != '' and v != ' '
            for v in [revenue, expenditures, investment, salary]
        )
        
        if not has_data and not month_val:
            empty_streak += 1
            if empty_streak >= MAX_EMPTY_STREAK:
                break
            continue
        else:
            empty_streak = 0
        
        if has_data and current_month:
            if row_data[0] and not any(v for v in row_data[1:]):
                continue
                
            entry_id += 1
            entries.append({
                'id': entry_id,
                'row': row_idx,
                'month': current_month,
                'month_num': current_month_num,
                'date': _parse_date(row_data[1]),
                'name': str(name).strip() if name else '',
                'type': str(row_data[3]).strip() if row_data[3] else '',
                'revenue': safe_float(revenue),
                'expenditures': safe_float(expenditures),
                'investment': safe_float(investment),
                'salary': safe_float(salary),
                'payment_method': str(row_data[8]).strip() if row_data[8] else '',
            })
    
    # Calculate monthly summary
    monthly_summary = {}
    for entry in entries:
        m = entry['month']
        if m not in monthly_summary:
            monthly_summary[m] = {
                'month': m,
                'month_num': entry['month_num'],
                'revenue': 0, 'expenditures': 0,
                'investment': 0, 'salary': 0
            }
        monthly_summary[m]['revenue'] += entry['revenue']
        monthly_summary[m]['expenditures'] += entry['expenditures']
        monthly_summary[m]['investment'] += entry['investment']
        monthly_summary[m]['salary'] += entry['salary']
    
    sorted_summary = dict(sorted(monthly_summary.items(), key=lambda x: x[1]['month_num']))
    
    # Store in memory cache
    result = (entries, sorted_summary)
    _cache[cache_key] = result
    _cache_mtime[cache_key] = current_mtime
    
    # Store in JSON cache for next time
    try:
        with open(json_cache_path, 'w', encoding='utf-8') as f:
            json.dump({
                'mtime': current_mtime,
                'entries': entries,
                'summary': sorted_summary
            }, f)
    except Exception as e:
        print("Failed to write JSON cache:", e)
        
    return result

def read_ledger(year_sheet='2026'):
    """
    Read all ledger entries from the Excel file.
    Returns a list of dicts and monthly summary.
    Uses in-memory cache to avoid re-reading the large Excel file.
    """
    cache_key = f'ledger_{year_sheet}'
    json_cache_path = os.path.join(os.path.dirname(__file__), f'{cache_key}.json')
    
    try:
        current_mtime = os.path.getmtime(EXCEL_PATH)
    except OSError:
        current_mtime = 0
        
    # 1. Check Memory Cache
    if cache_key in _cache and _cache_mtime.get(cache_key) == current_mtime:
        print("CACHE HIT (Memory)")
        return _cache[cache_key]
        
    # 2. Check JSON File Cache (Ultra Fast)
    try:
        if os.path.exists(json_cache_path):
            with open(json_cache_path, 'r', encoding='utf-8') as f:
                cached_data = json.load(f)
                if cached_data.get('mtime') == current_mtime:
                    print("CACHE HIT (JSON Disk)")
                    _cache[cache_key] = (cached_data['entries'], cached_data['summary'])
                    _cache_mtime[cache_key] = current_mtime
                    return _cache[cache_key]
    except Exception as e:
        print("JSON Cache error:", e)

    print(f"CACHE MISS! mtime: {current_mtime}. Reading Excel file...")
    t0 = time.time()
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    if year_sheet not in wb.sheetnames:
        wb.close()
        return [], {}
    
    result = _parse_and_cache(wb, year_sheet, current_mtime)
    wb.close()
    
    print(f"CACHE LOADED AND SAVED IN {time.time() - t0:.2f} seconds")
    return result


def get_month_boundaries(ws):
    """Get the start row for each month section in the sheet."""
    boundaries = {}
    max_scan = min(500, ws.max_row + 1)
    for row_idx in range(10, max_scan):
        val = ws.cell(row=row_idx, column=1).value
        month_num = _normalize_month(val)
        if month_num:
            boundaries[month_num] = row_idx
            empty_streak = 0
            if len(boundaries) == 12:
                break  # Found all months
        else:
            # Check if row has any data at all
            has_any = any(ws.cell(row=row_idx, column=c).value for c in range(2, 10))
            if not has_any:
                empty_streak += 1
                if empty_streak >= 50:
                    break
            else:
                empty_streak = 0
    return boundaries


def add_ledger_entry(month_num, date_str, name, entry_type, revenue, expenditures, investment, salary, payment_method, year_sheet='2026'):
    """
    Add a new entry to the Excel ledger under the correct month.
    Inserts a new row at the end of the month section (before the next month).
    """
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[year_sheet]
    
    boundaries = get_month_boundaries(ws)
    
    if month_num not in boundaries:
        wb.close()
        return False, "Month not found in ledger"
    
    # Find insert position: last data row of this month 
    # (before next month marker or end of data)
    month_start = boundaries[month_num]
    
    # Find next month start
    sorted_months = sorted(boundaries.keys())
    current_idx = sorted_months.index(month_num)
    
    if current_idx < len(sorted_months) - 1:
        next_month_start = boundaries[sorted_months[current_idx + 1]]
    else:
        # Last month - find end of data
        next_month_start = ws.max_row + 1
    
    # Find the last non-empty row in this month section
    insert_row = month_start + 1  # default: right after month marker
    for r in range(next_month_start - 1, month_start, -1):
        row_data = [ws.cell(row=r, column=c).value for c in range(1, 10)]
        if any(v is not None and v != '' and v != ' ' for v in row_data[1:]):
            insert_row = r + 1
            break
    
    # Insert a new row
    ws.insert_rows(insert_row)
    
    # Write data
    # Col A: empty (month marker only on first row)
    # Col B: Date
    if date_str:
        try:
            ws.cell(row=insert_row, column=2, value=datetime.strptime(date_str, '%Y-%m-%d'))
        except ValueError:
            ws.cell(row=insert_row, column=2, value=date_str)
    
    # Col C: Name
    ws.cell(row=insert_row, column=3, value=name)
    
    # Col D: Type
    ws.cell(row=insert_row, column=4, value=entry_type if entry_type else None)
    
    # Col E: Revenue
    if revenue:
        ws.cell(row=insert_row, column=5, value=float(revenue))
    
    # Col F: Expenditures
    if expenditures:
        ws.cell(row=insert_row, column=6, value=float(expenditures))
    
    # Col G: Investment
    if investment:
        ws.cell(row=insert_row, column=7, value=float(investment))
    
    # Col H: Salary
    if salary:
        ws.cell(row=insert_row, column=8, value=float(salary))
    
    # Col I: Payment Method
    ws.cell(row=insert_row, column=9, value=payment_method if payment_method else None)
    
    wb.save(EXCEL_PATH)
    
    # Update cache immediately so next page load is instant
    try:
        new_mtime = os.path.getmtime(EXCEL_PATH)
        _parse_and_cache(wb, year_sheet, new_mtime)
    except Exception as e:
        print("Error updating cache after add:", e)
        
    wb.close()
    
    return True, "Entry added successfully"


def update_ledger_entry(row_num, date_str, name, entry_type, revenue, expenditures, investment, salary, payment_method, year_sheet='2026'):
    """
    Update an existing entry in the Excel ledger by row number.
    """
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[year_sheet]
    
    # Update cells
    if date_str:
        try:
            ws.cell(row=row_num, column=2, value=datetime.strptime(date_str, '%Y-%m-%d'))
        except ValueError:
            ws.cell(row=row_num, column=2, value=date_str)
    else:
        ws.cell(row=row_num, column=2, value=None)
    
    ws.cell(row=row_num, column=3, value=name if name else None)
    ws.cell(row=row_num, column=4, value=entry_type if entry_type else None)
    
    def write_amount(col, val):
        if val and str(val).strip():
            try:
                ws.cell(row=row_num, column=col, value=float(val))
            except (ValueError, TypeError):
                ws.cell(row=row_num, column=col, value=None)
        else:
            ws.cell(row=row_num, column=col, value=None)
    
    write_amount(5, revenue)
    write_amount(6, expenditures)
    write_amount(7, investment)
    write_amount(8, salary)
    
    ws.cell(row=row_num, column=9, value=payment_method if payment_method else None)
    
    wb.save(EXCEL_PATH)
    
    # Update cache immediately
    try:
        new_mtime = os.path.getmtime(EXCEL_PATH)
        _parse_and_cache(wb, year_sheet, new_mtime)
    except Exception as e:
        print("Error updating cache after update:", e)
        
    wb.close()
    
    return True, "Entry updated successfully"

def delete_ledger_entry(row_num, year_sheet='2026'):
    """
    Delete an existing entry in the Excel ledger by row number.
    """
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[year_sheet]
    
    # Delete the entire row
    ws.delete_rows(row_num)
    
    wb.save(EXCEL_PATH)
    
    # Update cache immediately
    try:
        new_mtime = os.path.getmtime(EXCEL_PATH)
        _parse_and_cache(wb, year_sheet, new_mtime)
    except Exception as e:
        print("Error updating cache after delete:", e)
        
    wb.close()
    
    return True, "Entry deleted successfully"
